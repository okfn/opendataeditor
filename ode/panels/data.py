import json
import csv
import logging
from pathlib import Path
from frictionless import system

from PySide6.QtCore import Qt, QAbstractTableModel, QObject, Signal, Slot, QRunnable, QRect, QEvent
from PySide6.QtGui import QColor, QIcon, QKeyEvent
from PySide6.QtWidgets import QWidget, QVBoxLayout, QTableView, QLabel, QApplication, QStyledItemDelegate, QStyle

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import xlwt
import xlrd
from xlutils.copy import copy


from ode import utils
from ode.dialogs.metadata import ColumnMetadataDialog, ColumnMetadataField
from ode.file import File
from ode.shared import COLOR_RED, COLOR_BLUE
from ode.paths import Paths


DEFAULT_LIMIT_ERRORS = 1000

logger = logging.getLogger(__name__)


class DataWorkerSignals(QObject):
    """Define the signals for the DataWorker."""

    finished = Signal(tuple)
    messages = Signal(str)


class DataWorker(QRunnable):
    """Worker to execute all the reading and validation tasks.

    This worker will allow us to read and validate the data (that can take several
    seconds) in the background. By moving this logic to a Worker we can avoid the
    application to get freeze while reading and instead display proper messages to
    the user.
    """

    def __init__(self, filepath, sheet_name=None):
        super().__init__()
        self.file = File(filepath)
        self.signals = DataWorkerSignals()
        self.sheet_name = sheet_name
        self.resource = self.file.get_or_create_metadata(sheet_name).get("resource")

    @Slot()
    def run(self):
        """Reads and validates the data.

        We are using Resource.read_cells() because we want to read the data
        as is in the file in order to properly show data errors.

        We are using the system context to allow us to read from non-relative paths.

        This method emits a finished signal with all the data that the main UI requires to
        display the table and the errors.
        """
        with system.use_context(trusted=True):
            self.signals.messages.emit(QApplication.translate("DataWorker", "Reading file..."))
            data = self.resource.read_cells()
            self.signals.messages.emit(QApplication.translate("DataWorker", "Checking errors..."))
            report = self.resource.validate(limit_errors=DEFAULT_LIMIT_ERRORS)

        self.signals.messages.emit(QApplication.translate("DataWorker", "Drawing table..."))

        errors = []
        if not report.valid:
            try:
                # report.error is only available for single error report
                errors.append(report.error)
            except Exception:
                errors = report.tasks[0].errors

        self.signals.messages.emit(QApplication.translate("DataWorker", "Read and error checking finished."))
        self.signals.finished.emit((self.file.path, data, errors))


class ColumnMetadataIconDelegate(QStyledItemDelegate):
    """
    Custom delegate to render an icon in the first row of the table.
    """

    icon_clicked = Signal(object)

    def __init__(self, icon_path, parent=None):
        super().__init__(parent)
        self.icon_size = 14
        self.icon = QIcon(icon_path)

    def _get_icon_rect(self, option):
        """Get the rectangle where the icon should be painted."""
        return QRect(
            option.rect.right() - self.icon_size,
            option.rect.top(),
            self.icon_size,
            self.icon_size,
        )

    def paint(self, painter, option, index):
        """Paint the icon in the first row of the table adds blue background if mouse over."""
        super().paint(painter, option, index)
        if index.row() == 0:
            if option.state & QStyle.State_MouseOver:
                painter.fillRect(option.rect, COLOR_BLUE)
            self.icon.paint(painter, self._get_icon_rect(option))

    def editorEvent(self, event, model, option, index):
        """Handle mouse events for the first row."""
        if index.row() == 0 and event.type() == QEvent.MouseButtonPress:
            self.icon_clicked.emit(index.column())
            return True

        return super().editorEvent(event, model, option, index)


class FrictionlessTableModel(QAbstractTableModel):
    finished = Signal()

    def __init__(
        self,
        data=[],
        errors=[],
    ):
        super().__init__()
        self._data = data
        self._row_count = self._get_row_count()
        self.errors = self._get_errors(errors)
        self._column_count = self._get_column_count()

    def write_data(self, filepath: Path, sheet_name=None):
        """
        Write the data to a file in the format specified by the file extension.
        """
        extension = filepath.suffix.lower()
        if extension == ".csv":
            self.write_data_csv(filepath)
        elif extension == ".xlsx":
            self.write_data_xlsx(filepath, sheet_name)
        elif extension == ".xls":
            self.write_data_xls(str(filepath), sheet_name)
        else:
            raise ValueError(f"Unsupported format: {extension}. Use .csv, .xlsx or .xls")

    def write_data_csv(self, filepath: Path):
        """
        Write the data to a CSV file.
        """
        logger.info(f"Writing data to CSV file: {filepath}")
        resource = File(filepath).get_or_create_metadata().get("resource")
        dialect = resource.dialect.to_dict()
        csv_config = dialect.get("csv", None)

        # Default delimiter
        delimiter = ","
        if csv_config and "delimiter" in csv_config:
            delimiter = csv_config["delimiter"]

        with open(filepath, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile, delimiter=delimiter)
            # Write header
            writer.writerow(self._data[0])
            # Write data rows
            writer.writerows(self._data[1:])

        logger.info(f"Data saved in CSV format: {filepath}")

    def write_data_xlsx(self, filepath: Path, sheet_name=None):
        """
        Write the data to an Excel file.
        """
        logger.info(f"Writing data to Excel file: {filepath}")

        wb = load_workbook(filepath)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Deleting existing data in sheet: {sheet_name}")
            ws.delete_rows(1, ws.max_row)  # Delete all rows
        else:
            logger.error(f"Sheet {sheet_name} does not exist in the workbook: {filepath}")
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook: {filepath}")

        # Header row
        ws.append(self._data[0])

        # Data rows
        rows = self._data[1:]
        for row in rows:
            ws.append(row)

        wb.save(filepath)
        logger.info(f"Data saved in Excel format: {filepath}")

    def write_data_xls(self, filepath: str, sheet_name=None):
        """
        Write the data to an Excel XLS file.

        The filepath must be a string because xlwt does not support Path objects.
        """
        logger.info(f"Writing data to XLS file: {filepath}")

        wb = xlwt.Workbook()

        rb = xlrd.open_workbook(filepath, formatting_info=True)
        try:
            sheet_name_index = rb.sheet_names().index(sheet_name)
        except ValueError:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook: {filepath}")

        # We use xlutils to transform the xlrd book into an xlwt book
        # This will allow us to modify existing XLS files
        wb = copy(rb)
        ws = wb.get_sheet(sheet_name_index)

        for row_idx, row_data in enumerate(self._data):
            for col_idx, cell_value in enumerate(row_data):
                ws.write(row_idx, col_idx, cell_value)

        wb.save(filepath)
        logger.info(f"Data saved in XLS format: {filepath}")

    def write_error_xlsx(self, filepath: Path):
        """
        Write the errors to an Excel file in the specified directory
        painting with red the cells with errors.
        """
        wb = Workbook()
        data_sheet = wb.active
        data_sheet.title = self.tr("Data")
        errors_sheet = wb.create_sheet("Errors Description")
        errors_sheet.append(["Row", "Column", "Error Title", "Error Description"])

        blank_sheet = wb.create_sheet("Blank Rows")
        blank_sheet.append(["Row", "Error Description"])

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        errors_cells = list()
        blank_rows = set()

        for row_index, errors_in_row in enumerate(self.errors):
            if errors_in_row:
                for error_column, error_type, error_description in errors_in_row:
                    if error_type == "blank-row":
                        blank_rows.add(row_index)
                    else:
                        errors_cells.append((row_index, error_column, error_type, error_description))

        for row_index, row in enumerate(self._data):
            data_sheet.append(row)

        errors_cells.sort(key=lambda x: (x[0], x[1]))  # Sort by row and column index
        for row_index, col_index, error_type, error_description in errors_cells:
            excel_row = row_index + 1
            excel_col = col_index + 1
            error_title = utils.ErrorTexts.get_error_title(error_type)
            if not error_title:
                error_title = error_type.replace("-", " ").title()

            errors_sheet.append(
                [
                    excel_row,
                    excel_col,
                    error_title,
                    utils.ErrorTexts.get_error_description(error_type) or error_description,
                ]
            )

            # Paint the cell with red if it has an error in the Data Sheet
            data_sheet.cell(row=excel_row, column=excel_col).fill = red_fill

        for row_index in blank_rows:
            excel_row = row_index + 1

            blank_sheet.append(
                [
                    excel_row,
                    utils.ErrorTexts.get_error_description("blank-row"),
                ]
            )

            # Paint the cell with red if it has an error in the Data Sheet
            for col_index in range(1, data_sheet.max_column + 1):
                data_sheet.cell(row=excel_row, column=col_index).fill = red_fill

        wb.save(filepath)
        self.finished.emit()
        logger.info(f"Errors saved in Excel format: {filepath}")

    def _get_errors(self, errors):
        """Return an array with errors information to use when rendering the table.

        The array has same lenght as our data with None or a Tuple:
          [None, None, (3, 'blank-label', 'error mesage to be displayed'), None]

        Main actions:
         - Moves from Frictionless' 1-index to 0-index
         - Handles inconsistency in Fricionless Error Object API
         - Builds an array for easy access to error information to render performant tables.
        """
        result = [None] * self._row_count
        for error in errors:
            if error.type == "source-error":
                # SourceError happens with files that cannot be read and do not have row_number nor field_number.
                return result
            elif error.type in ["blank-label", "duplicate-label", "incorrect-label", "missing-label", "extra-label"]:
                # https://github.com/frictionlessdata/frictionless-py/issues/1710
                row = error.row_numbers[0] - 1
                column = error.field_number - 1
            elif error.type == "blank-row":
                row = error.row_number - 1
                # BlankRow error does not have field_number
                column = 0
            elif not hasattr(error, "row_number"):
                row = None
                column = None
            else:
                row = error.row_number - 1
                column = error.field_number - 1

            if row is not None:
                if result[row] is None:
                    result[row] = list()

                result[row].append((column, error.type, error.message))

        return result

    def _get_row_count(self):
        return len(self._data)

    def _get_column_count(self):
        """Get the amout of columns.

        We are expecting malformed CSVs, so the amout of columns should always
        be the size of the longest row.
        """
        try:
            return max(map(len, self._data))
        except ValueError:
            return 0

    def get_header_data(self):
        """Returns the first row of the file."""
        return self._data[0]

    def rowCount(self, parent=None):
        """Returning from a pre-calculated private attribute for performance improvements."""
        return self._row_count

    def columnCount(self, parent=None):
        """Returning from a pre-calculated private attribute for performance improvements."""
        return self._column_count

    def data(self, index, role):
        """Returns information to be used to render the Data Table View.

        For each cell we return:
         - The value to be displayed in the cell
         - If there is an error in that cell, a red color for the Background.
         - If there is an error in that cell, a message for the tooltip.
        """
        if not index.isValid():
            return None

        if role == Qt.ItemDataRole.DisplayRole or role == Qt.ItemDataRole.EditRole:
            try:
                # We convert the data as string to avoid PySide6 treating numbers differently
                value = self._data[index.row()][index.column()]
                if value is None:
                    return ""
                return str(value)
            except IndexError:
                # Our data could be irregular (missing columns and rows)
                # So it is okay to return None and keep iterating.
                return None

        if not self.errors[index.row()]:
            return None

        if role == Qt.ItemDataRole.BackgroundRole:
            for error_column, error_type, _ in self.errors[index.row()]:
                if error_type == "blank-row":
                    # BlankRowError does not have field_number, we paint all the cells.
                    return COLOR_RED
                if error_column == index.column():
                    return COLOR_RED
        elif role == Qt.ForegroundRole:
            for error_column, _, _ in self.errors[index.row()]:
                if error_column == index.column():
                    return QColor(255, 255, 255)

    def flags(self, index):
        """Enable edition mode"""
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags

        if index.row() == 0:
            return super().flags(index)

        return super().flags(index) | Qt.ItemFlag.ItemIsEditable

    def setData(self, index, value, role):
        """Insert the edited value at the specific cell.

        Since the TableView will always have enough rows and enough columns,
        when editing the raw data we encounter two scenarios:
          a) The cell in the raw data exist and therefore we just replace
          b) The cell in the raw data do not exist and therefore we need to
          create empty cells until we get to the exact column we want to insert.
        """
        if role == Qt.ItemDataRole.EditRole:
            currentRow = self._data[index.row()]
            try:
                # a) raw cell exist
                currentRow[index.column()] = value
            except IndexError:
                # b) we create empty cells and then insert at specific column
                currentRow += [None] * (index.column() - len(currentRow))
                currentRow.insert(index.column(), value)
            self.dataChanged.emit(index, index)
            return True
        return False


class CustomTableView(QTableView):
    """Custom QTableView to handle specific key events."""

    on_click_first_row = Signal(object)

    def __init__(self, parent=None):
        super().__init__(parent)

    def keyPressEvent(self, event: QKeyEvent):
        """
        We check if the user pressed Enter or Return on the first row of the table.
        If so, we emit a signal with the column index of the clicked cell.
        """
        index = self.currentIndex()

        if event.key() in (Qt.Key_Return, Qt.Key_Enter) and index.row() == 0 and index.isValid():
            self.on_click_first_row.emit(index.column())
            return

        super().keyPressEvent(event)

    def mouseMoveEvent(self, event):
        """Changes the cursor to Pointing Hand if positing is in first row."""
        index = self.indexAt(event.pos())
        if index.row() == 0:
            self.setCursor(Qt.PointingHandCursor)
        else:
            self.unsetCursor()
        return super().mouseMoveEvent(event)


class DataViewer(QWidget):
    """Widget to display the content of tabular data."""

    # Signal to notify that the metadata has been saved
    on_save = Signal(object)

    def __init__(self):
        super().__init__()

        utils.set_common_style(self)

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.label = QLabel()
        self.label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)

        # Create the frozen table widget but hide it until we need it
        self.frozen_table_widget = FrozenRowTableWidget()
        self.frozen_table_widget.on_click_first_row.connect(self.show_column_metadata_dialog)
        self.frozen_table_widget.hide()

        # Keep the reference to the main table view
        self.table_view = self.frozen_table_widget.main_table

        self.delegate = ColumnMetadataIconDelegate(Paths.asset("icons/three-lines.png"))
        self.delegate.icon_clicked.connect(self.show_column_metadata_dialog)

        layout.addWidget(self.label)
        layout.addWidget(self.frozen_table_widget)

        self.retranslateUI()

    def display_data(self, model, filepath, sheet_name=None):
        """Set the model of the QTableView

        When a tabular file is selected, the main application will create a
        FrictionlessTableModel and call this function using the model as a parametner.
        """
        self.frozen_table_widget.setModel(model)

        self.frozen_table_widget.frozen_table.setItemDelegate(self.delegate)

        self.frozen_table_widget.frozen_table.horizontalHeader().setDefaultSectionSize(120)
        self.table_view.horizontalHeader().setDefaultSectionSize(120)
        self.table_view.setMouseTracking(True)

        self.metadata = File(filepath).get_or_create_metadata(sheet_name)
        self.resource = self.metadata.get("resource")

        self.label.hide()
        self.frozen_table_widget.show()

    def show_column_metadata_dialog(self, field_index):
        """
        Shows a dialog to edit a column's metadata.
        """
        model = self.frozen_table_widget.original_model
        column_count = model.columnCount()
        headers = []
        for column in range(column_count):
            index = model.index(0, column)
            value = model.data(index, Qt.DisplayRole)
            headers.append(value)

        field = self.resource.schema.fields[field_index]
        column_metadata_field = ColumnMetadataField(
            headers[field_index], field.type, field.description, field.constraints
        )
        dialog = ColumnMetadataDialog(self, column_metadata_field, field_index, headers)
        dialog.save_clicked.connect(self.save_metadata_to_descriptor_file)
        dialog.exec()

    def clear(self, model):
        """Reset the view to the default state.

        This view depends of the main application self.table_model attribute. This
        method should always receive an empty model
        """
        self.frozen_table_widget.setModel(model)
        self.label.show()
        self.frozen_table_widget.hide()

    def retranslateUI(self):
        """Apply translations to class elements."""
        self.label.setText(self.tr("Preview not available for this item."))

    def save_metadata_to_descriptor_file(self, field_form: dict):
        """Save the metadata to the descriptor file."""
        field_index = field_form.get("index")
        field = self.resource.schema.fields[field_index]

        field.name = field_form.get("name")
        field.title = field_form.get("name")
        field.description = field_form.get("description", "")
        field.constraints = {
            "required": field_form.get("constraints").get("required"),
        }

        type = field_form.get("type")

        if type == "string":
            field.constraints["minLength"] = field_form.get("constraints").get("minLength")
            field.constraints["maxLength"] = field_form.get("constraints").get("maxLength")
        else:
            # If the type is not Text, we remove the minLength and maxLength constraints
            field.constraints.pop("minLength", None)
            field.constraints.pop("maxLength", None)

        # Update the field in the schema
        self.resource.schema.set_field(field)

        # Field.type cannot be updated directly, we need to use set_field_type
        # it needs to be after the set_field to avoid being overridden
        self.resource.schema.set_field_type(field.name, type)

        self.metadata["resource"] = self.resource.to_descriptor()
        file = File(self.resource.path)

        # We remove the dialect from the metatada, because Frictionless will be stuck
        # on this sheet otherwise
        self.metadata["resource"].pop("dialect", None)

        with open(file.metadata_path, "w") as f:
            print(f"Saving metadata {file.metadata_path}")
            json.dump(self.metadata, f)

        # Check if we name was changed, if so we need to update the header
        model = self.frozen_table_widget.original_model
        index = model.index(0, field_index)
        original_name = model.data(index, Qt.DisplayRole)

        table_view_changed = False
        if original_name != field.name:
            model.setData(index, field.name, Qt.EditRole)
            table_view_changed = True

        self.on_save.emit(table_view_changed)


class FrozenRowTableWidget(QWidget):
    """Widget that contains two QTableView to simulate a frozen first row."""

    on_click_first_row = Signal(object)

    def __init__(self, parent=None):
        super().__init__(parent)

        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)  # No margins
        self.layout.setSpacing(0)  # No spacing between widgets
        self.setLayout(self.layout)

        # Table for the frozen first row
        self.frozen_table = CustomTableView()
        self.frozen_table.setMaximumHeight(56)  # Row height
        self.frozen_table.setMinimumHeight(56)
        self.frozen_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.frozen_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.frozen_table.setSelectionBehavior(QTableView.SelectColumns)
        self.frozen_table.setCornerButtonEnabled(False)
        self.frozen_table.setStyleSheet("QTableView { border-bottom: none; }")

        # Table for the rest of the data
        self.main_table = CustomTableView()
        self.main_table.setCornerButtonEnabled(False)
        self.main_table.setTabKeyNavigation(False)
        self.main_table.setStyleSheet("QTableView { border-top: none;border-left:1px solid #dcdcdc;}")

        # Synchronize horizontal scrollbars
        self.frozen_table.horizontalScrollBar().valueChanged.connect(self.main_table.horizontalScrollBar().setValue)
        self.main_table.horizontalScrollBar().valueChanged.connect(self.frozen_table.horizontalScrollBar().setValue)

        # Hook to handle clicks on the first row
        self.frozen_table.on_click_first_row.connect(self._on_frozen_row_clicked)

        # Add both tables to the layout
        self.layout.addWidget(self.frozen_table)
        self.layout.addWidget(self.main_table)

        # Variable to keep the reference to the original model
        self.original_model = None

    def setModel(self, model):
        """Set the model for both tables."""
        self.original_model = model

        # Model for the frozen first row
        frozen_model = FrozenRowModel(model)
        self.frozen_table.setModel(frozen_model)

        # Model for the main data (excluding the first row)
        main_model = MainDataModel(model)
        self.main_table.setModel(main_model)

        # Synchronize column widths
        self._sync_column_widths()

        # Horizontal headers status
        self.frozen_table.horizontalHeader().setVisible(True)
        self.main_table.horizontalHeader().setVisible(False)

        # Vertical headers status
        self.frozen_table.verticalHeader().setVisible(True)
        self.main_table.verticalHeader().setVisible(True)

        self.frozen_table.verticalHeader().setDefaultSectionSize(self.main_table.verticalHeader().defaultSectionSize())
        self.frozen_table.verticalHeader().setFixedWidth(self.main_table.verticalHeader().width())

    def _sync_column_widths(self):
        """
        Synchronize column widths between the frozen and main tables.
        """
        frozen_header = self.frozen_table.horizontalHeader()
        main_header = self.main_table.horizontalHeader()

        # Hook to resizeSection to avoid infinite loop
        frozen_header.sectionResized.connect(lambda idx, old_size, new_size: main_header.resizeSection(idx, new_size))
        main_header.sectionResized.connect(lambda idx, old_size, new_size: frozen_header.resizeSection(idx, new_size))

        # Set initial sizes
        column_count = self.original_model.columnCount() if self.original_model else 0
        for i in range(column_count):
            frozen_header.resizeSection(i, 120)
            main_header.resizeSection(i, 120)

    def _on_frozen_row_clicked(self, column_index):
        """Emit the signal when a column in the frozen row is clicked."""
        self.on_click_first_row.emit(column_index)


class FrozenRowModel(QAbstractTableModel):
    """Model to show only the first row of the original model"""

    def __init__(self, original_model):
        super().__init__()
        self.original_model = original_model

    def rowCount(self, parent=None):
        return 1  # Return only one row

    def columnCount(self, parent=None):
        return self.original_model.columnCount() if self.original_model else 0

    def data(self, index, role):
        if not index.isValid() or not self.original_model:
            return None

        # Always get data from the first row of the original model
        original_index = self.original_model.index(0, index.column())
        return self.original_model.data(original_index, role)

    def setData(self, index, value, role):
        if not index.isValid() or not self.original_model:
            return False

        # Edit directly in the original model
        original_index = self.original_model.index(0, index.column())
        return self.original_model.setData(original_index, value, role)

    def flags(self, index):
        if not index.isValid() or not self.original_model:
            return Qt.NoItemFlags

        original_index = self.original_model.index(0, index.column())
        return self.original_model.flags(original_index)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        """Setup custom headers for the frozen row model"""
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return str(section + 1)
            elif orientation == Qt.Vertical:
                # Always 1 because we only have one row
                return "1"
        return None


class MainDataModel(QAbstractTableModel):
    """
    Model to show all the rows except the first one of the original model
    """

    def __init__(self, original_model):
        super().__init__()
        self.original_model = original_model

    def rowCount(self, parent=None):
        if not self.original_model:
            return 0
        # All the rows except the first one
        return max(0, self.original_model.rowCount() - 1)

    def columnCount(self, parent=None):
        return self.original_model.columnCount() if self.original_model else 0

    def data(self, index, role):
        if not index.isValid() or not self.original_model:
            return None

        # Mapping the row index to the original model (offset by 1)
        original_index = self.original_model.index(index.row() + 1, index.column())
        return self.original_model.data(original_index, role)

    def setData(self, index, value, role):
        if not index.isValid() or not self.original_model:
            return False

        # We edit directly in the original model (offset by 1)
        original_index = self.original_model.index(index.row() + 1, index.column())
        result = self.original_model.setData(original_index, value, role)

        if result:
            self.dataChanged.emit(index, index)

        return result

    def flags(self, index):
        if not index.isValid() or not self.original_model:
            return Qt.NoItemFlags

        original_index = self.original_model.index(index.row() + 1, index.column())
        return self.original_model.flags(original_index)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        """Setup custom headers for the main data model"""
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return str(section + 1)
            elif orientation == Qt.Vertical:
                # Offset by 1 because we do not show the first row
                return str(section + 2)
        return None
